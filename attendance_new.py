import datetime
from openpyxl import load_workbook
import pandas as pd
import os

class Register:


    #constructor invoked when the object is created. it initializes the required variable
    def __init__(self,empid,current_date,current_time):
        self.empid=empid
        self.current_day=datetime.datetime.now().strftime("%A")
        self.current_date=current_date
        self.current_time=current_time
        self.idx=[]
        self.filename=os.getcwd()+'/attendence_sheet.xlsx'


    #the function to load the excel sheet for entry,update or edit
    def load_excel_sheet(self):
        wb = load_workbook(self.filename)
        ws = wb.worksheets[0]
        return wb,ws

    #the function just keeps the arguments in the form of the list so as to make it easy to append to the  file using openpyxl
    def format_data(self):
        return ([self.empid,self.current_date,self.current_day,self.current_time])

    #It checks the existing excel file to check whether the employee is already registered or not. If not present it will send false and if present it sends true
    def check_excel(self):
        df = pd.read_excel(self.filename)
        status=((df['Emp_ID'] == self.empid) & (df['Date'] == self.current_date)).any()
        print(status)
        return status

    #The function that insert the user based on the different conditions
    def insert_user_attendance(self,data):
        # print(data)
        if data[3]<=(self.current_time.replace(hour=12,minute=0,second=0)):
            if self.check_excel()==False:
                wb,ws=self.load_excel_sheet()
                ws.append(data)
                wb.save(self.filename)
                msg="Thank you"


            elif self.check_excel()== True:
                msg="you have already registered"
                print(msg)

        elif data[3]>self.current_time.replace(hour=12,minute=0,second=0):
            if self.check_excel() == False:
                state=0
                self.pd_write_to_excel_sheet(self.empid,self.current_date,self.current_time,state)

            elif self.check_excel()==True:
                state=1
                self.pd_write_to_excel_sheet(self.empid,self.current_date,self.current_time,state)
        return 0
    #function that updates the exit time on the existing spread sheet
    def pd_write_to_excel_sheet(self,empid,date,time,status):
        entry=[empid,date,self.current_day]
        df=pd.read_excel(self.filename)
        wb,ws=self.load_excel_sheet()
        if status==0:
            print("iam in the false state")
            adent=["Notavailable",self.current_time]
            entry.extend(adent)
            ws.append(entry)
            wb.save(self.filename)
            print("sucessfull")
        elif status==1:
            print("I am here")
            print(df)
            index=df.index[(df['Emp_ID']== self.empid) & (df['Date']==self.current_date)].tolist()
            #x = df[df['Emp_ID'] == empid & (df['Date'] == date)].index.values.astype(int)
            print(index)
            ws.cell(row=int(index[0]+2), column=5).value = time
            wb.save(self.filename)



def attendance_register(empid):
    reg=Register(empid,datetime.datetime.now().date(), datetime.datetime.now().time())
    # reg=Register(empid)
    reg.insert_user_attendance(reg.format_data())
    return 0
